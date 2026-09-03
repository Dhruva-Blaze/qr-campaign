import os
import base64
import uuid
import threading
from datetime import datetime, timezone, timedelta
from functools import wraps
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import certifi
import jwt
from bson import ObjectId
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from gridfs import GridFS
from pymongo import MongoClient, DESCENDING
from werkzeug.security import check_password_hash

app = Flask(__name__)

MONGO_URI = os.environ.get("MONGO_URI", "").strip()
DB_NAME = os.environ.get("MONGO_DB_NAME", "photo_portal").strip() or "photo_portal"
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
JWT_SECRET = os.environ.get("JWT_SECRET", "")

if not MONGO_URI or not ADMIN_PASSWORD or not JWT_SECRET:
    raise RuntimeError("Set MONGO_URI, ADMIN_PASSWORD and JWT_SECRET environment variables.")

raw_origins = os.environ.get("FRONTEND_ORIGINS", "").strip()
origins = [item.strip().rstrip("/") for item in raw_origins.split(",") if item.strip()]
if not origins:
    origins = ["https://gl.blazecorporation.in", "http://localhost", "http://127.0.0.1"]

CORS(
    app,
    resources={r"/api/*": {"origins": origins}},
    methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

_mongo_lock = threading.Lock()
_client = None
_db = None
_inquiries = None
_fs = None


def get_mongo():
    """Create and verify MongoClient lazily inside the running worker."""
    global _client, _db, _inquiries, _fs

    with _mongo_lock:
        if _client is not None:
            try:
                _client.admin.command("ping")
                return _client, _db, _inquiries, _fs
            except Exception:
                try:
                    _client.close()
                except Exception:
                    pass
                _client = _db = _inquiries = _fs = None

        options = {
            "serverSelectionTimeoutMS": 30000,
            "connectTimeoutMS": 30000,
            "socketTimeoutMS": 30000,
            "retryWrites": True,
            "tls": True,
            "tlsCAFile": certifi.where(),
            "tlsDisableOCSPEndpointCheck": True,
            "appname": "qr-campaign",
        }

        client = MongoClient(MONGO_URI, **options)
        try:
            client.admin.command("ping")
        except Exception:
            client.close()
            raise

        _client = client
        _db = client[DB_NAME]
        _inquiries = _db.inquiries
        _fs = GridFS(_db)
        return _client, _db, _inquiries, _fs


def verify_password(password):
    if ADMIN_PASSWORD.startswith(("pbkdf2:", "scrypt:")):
        return check_password_hash(ADMIN_PASSWORD, password)
    return password == ADMIN_PASSWORD


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        header = request.headers.get("Authorization", "")
        if not header.startswith("Bearer "):
            return jsonify(message="Unauthorized"), 401
        try:
            jwt.decode(header[7:], JWT_SECRET, algorithms=["HS256"])
        except Exception:
            return jsonify(message="Unauthorized"), 401
        return fn(*args, **kwargs)

    return wrapper


def serialize(doc):
    return {
        "id": str(doc["_id"]),
        "name": doc.get("name", ""),
        "mobile": doc.get("mobile", ""),
        "city": doc.get("city", ""),
        "looking_for_home_loan": doc.get("looking_for_home_loan", ""),
        "loan_purpose": doc.get("loan_purpose", ""),
        "loan_amount": doc.get("loan_amount", ""),
        "occupation": doc.get("occupation", ""),
        "selected_frame": doc.get("selected_frame", ""),
        "created_at": doc.get("created_at"),
        "has_photo": bool(doc.get("photo_file_id")),
    }


@app.errorhandler(404)
def not_found(_error):
    return jsonify(message="Not found"), 404


@app.errorhandler(Exception)
def handle_error(error):
    app.logger.exception(error)
    return jsonify(message=f"Server error: {str(error)}"), 500


@app.get("/")
def root():
    return jsonify(service="photo-portal-api", status="ok")


@app.get("/api/health")
def health():
    try:
        client, _, _, _ = get_mongo()
        client.admin.command("ping")
        return jsonify(status="ok", database="connected", database_name=DB_NAME)
    except Exception as error:
        return jsonify(status="error", message=f"MongoDB connection failed: {error}"), 503


@app.post("/api/inquiries")
def create_inquiry():
    data = request.get_json(silent=True) or {}

    required = ["name", "mobile", "city", "looking_for_home_loan", "image"]
    missing = [key for key in required if not str(data.get(key, "")).strip()]
    if missing:
        return jsonify(message="Missing required fields: " + ", ".join(missing)), 400

    loan_choice = str(data.get("looking_for_home_loan", "")).strip()
    if loan_choice not in ("Yes", "No"):
        return jsonify(message="Invalid home loan choice"), 400

    if loan_choice == "Yes":
        loan_required = ["loan_purpose", "loan_amount", "occupation"]
        missing_loan = [
            key for key in loan_required if not str(data.get(key, "")).strip()
        ]
        if missing_loan:
            return jsonify(
                message="Complete all home loan fields: " + ", ".join(missing_loan)
            ), 400

    image = str(data.get("image", ""))
    try:
        payload = image.split(",", 1)[1] if "," in image else image
        raw = base64.b64decode(payload, validate=True)
    except Exception:
        return jsonify(message="Invalid image data"), 400

    if len(raw) > 12 * 1024 * 1024:
        return jsonify(message="Image is too large"), 400

    try:
        _, _, inquiries, fs = get_mongo()

        file_id = fs.put(
            raw,
            filename=f"{uuid.uuid4().hex}.png",
            content_type="image/png",
        )

        doc = {
            "name": str(data.get("name", "")).strip(),
            "mobile": str(data.get("mobile", "")).strip(),
            "city": str(data.get("city", "")).strip(),
            "looking_for_home_loan": loan_choice,
            "loan_purpose": str(data.get("loan_purpose", "")).strip()
            if loan_choice == "Yes"
            else "",
            "loan_amount": str(data.get("loan_amount", "")).strip()
            if loan_choice == "Yes"
            else "",
            "occupation": str(data.get("occupation", "")).strip()
            if loan_choice == "Yes"
            else "",
            "selected_frame": str(data.get("selected_frame", "")).strip(),
            "photo_file_id": file_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

        try:
            result = inquiries.insert_one(doc)
        except Exception:
            try:
                fs.delete(file_id)
            except Exception:
                pass
            raise

        return jsonify(success=True, id=str(result.inserted_id)), 201

    except Exception as error:
        app.logger.exception(error)
        return jsonify(message=f"Could not save inquiry: {str(error)}"), 503


@app.post("/api/admin/login")
def login():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username", ""))
    password = str(data.get("password", ""))

    if username != ADMIN_USERNAME or not verify_password(password):
        return jsonify(message="Invalid username or password"), 401

    now = datetime.now(timezone.utc)
    token = jwt.encode(
        {"sub": ADMIN_USERNAME, "iat": now, "exp": now + timedelta(hours=12)},
        JWT_SECRET,
        algorithm="HS256",
    )
    return jsonify(token=token)


@app.get("/api/admin/stats")
@admin_required
def stats():
    _, _, inquiries, _ = get_mongo()
    today = datetime.now(timezone.utc).date().isoformat()
    return jsonify(
        total_inquiries=inquiries.count_documents({}),
        today_inquiries=inquiries.count_documents(
            {"created_at": {"$regex": "^" + today}}
        ),
        total_photos=inquiries.count_documents(
            {"photo_file_id": {"$exists": True}}
        ),
    )


@app.get("/api/admin/inquiries")
@admin_required
def list_inquiries():
    _, _, inquiries, _ = get_mongo()
    return jsonify(
        [serialize(item) for item in inquiries.find().sort("created_at", DESCENDING)]
    )


@app.get("/api/admin/export/leads.xlsx")
@admin_required
def export_leads_excel():
    """Download all received inquiry leads as an Excel workbook."""
    _, _, inquiries, _ = get_mongo()
    records = list(inquiries.find().sort("created_at", DESCENDING))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    headers = [
        "Lead ID",
        "Name",
        "Mobile",
        "City",
        "Looking for Home Loan",
        "Loan Purpose",
        "Loan Amount",
        "Occupation",
        "Selected Frame",
        "Received At",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="0083CA")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for doc in records:
        sheet.append([
            str(doc.get("_id", "")),
            doc.get("name", ""),
            doc.get("mobile", ""),
            doc.get("city", ""),
            doc.get("looking_for_home_loan", ""),
            doc.get("loan_purpose", ""),
            doc.get("loan_amount", ""),
            doc.get("occupation", ""),
            doc.get("selected_frame", ""),
            doc.get("created_at", ""),
        ])

    sheet.freeze_panes = "A2"
    for column_index, header in enumerate(headers, start=1):
        max_length = len(header)
        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            value = row[0].value
            max_length = max(max_length, len(str(value or "")))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 32)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"received-leads-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.get("/api/admin/photos/<inquiry_id>")
@admin_required
def photo(inquiry_id):
    _, _, inquiries, fs = get_mongo()
    try:
        doc = inquiries.find_one({"_id": ObjectId(inquiry_id)})
    except Exception:
        doc = None

    if not doc or not doc.get("photo_file_id"):
        return jsonify(message="Not found"), 404

    file = fs.get(doc["photo_file_id"])
    return send_file(
        BytesIO(file.read()),
        mimetype=file.content_type or "image/png",
    )


@app.delete("/api/admin/inquiries/<inquiry_id>")
@admin_required
def delete(inquiry_id):
    _, _, inquiries, fs = get_mongo()
    try:
        doc = inquiries.find_one({"_id": ObjectId(inquiry_id)})
    except Exception:
        doc = None

    if not doc:
        return jsonify(message="Not found"), 404

    if doc.get("photo_file_id"):
        fs.delete(doc["photo_file_id"])

    inquiries.delete_one({"_id": doc["_id"]})
    return jsonify(success=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
